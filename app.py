"""
VENTAS DE COMBUSTIBLE Y TIENDA FULL — FARMEX
Todo en un solo archivo para evitar problemas de despliegue.
Desarrollado por Lucas Sellecchia.
"""
from __future__ import annotations

import re
import sqlite3
import uuid
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook

# ============================================================================
# STORAGE
# ============================================================================

DB_PATH = Path(__file__).parent / "estacion_data.db"


def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS combustible (
            id TEXT PRIMARY KEY,
            fecha TEXT NOT NULL,
            turno TEXT NOT NULL,
            producto TEXT NOT NULL,
            litros REAL NOT NULL,
            importe REAL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS tienda (
            id TEXT PRIMARY KEY,
            fecha TEXT NOT NULL,
            turno TEXT NOT NULL,
            categoria TEXT NOT NULL,
            importe REAL DEFAULT 0,
            cantidad REAL DEFAULT 0
        )
        """
    )
    conn.commit()
    return conn


def load_combustible() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM combustible", conn)
    conn.close()
    return df


def load_tienda() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM tienda", conn)
    conn.close()
    return df


def insert_combustible(rows: list[dict]) -> tuple[int, int]:
    """Inserta filas nuevas, ignorando duplicados por (fecha, turno, producto)."""
    if not rows:
        return 0, 0
    existing = load_combustible()
    existing_keys = (
        set(zip(existing["fecha"], existing["turno"], existing["producto"]))
        if not existing.empty
        else set()
    )
    conn = get_conn()
    cur = conn.cursor()
    added, skipped, seen = 0, 0, set()
    for r in rows:
        key = (r["fecha"], r["turno"], r["producto"])
        if key in existing_keys or key in seen:
            skipped += 1
            continue
        seen.add(key)
        cur.execute(
            "INSERT INTO combustible (id, fecha, turno, producto, litros, importe) VALUES (?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), r["fecha"], r["turno"], r["producto"], r["litros"], r.get("importe", 0)),
        )
        added += 1
    conn.commit()
    conn.close()
    return added, skipped


def insert_tienda(rows: list[dict]) -> tuple[int, int]:
    """Inserta filas nuevas, ignorando duplicados por (fecha, turno, categoria)."""
    if not rows:
        return 0, 0
    existing = load_tienda()
    existing_keys = (
        set(zip(existing["fecha"], existing["turno"], existing["categoria"]))
        if not existing.empty
        else set()
    )
    conn = get_conn()
    cur = conn.cursor()
    added, skipped, seen = 0, 0, set()
    for r in rows:
        key = (r["fecha"], r["turno"], r["categoria"])
        if key in existing_keys or key in seen:
            skipped += 1
            continue
        seen.add(key)
        cur.execute(
            "INSERT INTO tienda (id, fecha, turno, categoria, importe, cantidad) VALUES (?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), r["fecha"], r["turno"], r["categoria"], r.get("importe", 0), r.get("cantidad", 0)),
        )
        added += 1
    conn.commit()
    conn.close()
    return added, skipped


def delete_combustible_ids(ids: list[str]):
    if not ids:
        return
    conn = get_conn()
    conn.executemany("DELETE FROM combustible WHERE id = ?", [(i,) for i in ids])
    conn.commit()
    conn.close()


def delete_tienda_ids(ids: list[str]):
    if not ids:
        return
    conn = get_conn()
    conn.executemany("DELETE FROM tienda WHERE id = ?", [(i,) for i in ids])
    conn.commit()
    conn.close()


# ============================================================================
# PARSERS
# ============================================================================

# ---------------------------------------------------------------------------
# Combustible
# ---------------------------------------------------------------------------

RAW_PRODUCT_MAP = {
    "NAFTA SUPER": "Nafta Súper",
    "DIESEL 500": "Diesel 500",
    "INFINIA NAFTA": "Infinia",
    "INFINIA DIESEL": "Infinia Diesel",
}
RAW_TURNO_MAP = {"1": "Noche", "2": "Mañana", "3": "Tarde"}
RAW_DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})\s*\((\d)\)")

PRODUCTOS = ["Nafta Súper", "Infinia", "Diesel 500", "Infinia Diesel"]
TURNOS = ["Mañana", "Tarde", "Noche"]


def parse_combustible_xlsx(file) -> list[dict]:
    """Lee la planilla cruda del aforador (mismo formato que siempre subís)."""
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c or "").strip().lower() == "fecha apertura" for c in row):
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = [str(c or "").strip().upper() for c in rows[header_idx]]
    col_idx = {key: headers.index(key) for key in RAW_PRODUCT_MAP if key in headers}

    result = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None for c in row):
            continue
        m = RAW_DATE_RE.search(str(row[0] or ""))
        if not m:
            continue
        dd, mm, yyyy, turno_n = m.groups()
        fecha = f"{yyyy}-{mm}-{dd}"
        turno = RAW_TURNO_MAP.get(turno_n, "Mañana")
        for raw_col, producto in RAW_PRODUCT_MAP.items():
            idx = col_idx.get(raw_col)
            if idx is None or idx >= len(row):
                continue
            try:
                val = float(row[idx])
            except (TypeError, ValueError):
                continue
            if not val:
                continue
            result.append({"fecha": fecha, "turno": turno, "producto": producto, "litros": val, "importe": 0})
    return result


def parse_combustible_csv(file) -> tuple[list[dict], int]:
    df = pd.read_csv(file)
    df.columns = [c.strip().lower() for c in df.columns]
    turnos_validos = {t.lower(): t for t in TURNOS}
    productos_validos = {p.lower(): p for p in PRODUCTOS}
    result, errors = [], 0
    for _, r in df.iterrows():
        fecha = str(r.get("fecha", "")).strip()
        turno = turnos_validos.get(str(r.get("turno", "")).strip().lower())
        producto = productos_validos.get(str(r.get("producto", "")).strip().lower())
        try:
            litros = float(r.get("litros"))
        except (TypeError, ValueError):
            errors += 1
            continue
        try:
            importe = float(r.get("importe", 0) or 0)
        except (TypeError, ValueError):
            importe = 0
        if not fecha or not turno or not producto:
            errors += 1
            continue
        result.append({"fecha": fecha, "turno": turno, "producto": producto, "litros": litros, "importe": importe})
    return result, errors


# ---------------------------------------------------------------------------
# Tienda Full
# ---------------------------------------------------------------------------

TURNOS_TIENDA = ["Mañana", "Tarde"]


def normalize_tienda_turno(raw: str):
    s = re.sub(r"[.\s]", "", str(raw or "").strip().lower())
    if s == "tm" or s.startswith("mañ") or s.startswith("man"):
        return "Mañana"
    if s == "tt" or s.startswith("tar"):
        return "Tarde"
    return None


def categoria_grupo(categoria: str):
    c = categoria or ""
    if re.match(r"^comidas?\s*(envasad|elaborad)", c, re.I):
        return "comida"
    if re.match(r"^bebidas?\s*calient", c, re.I):
        return "cafe"
    if re.match(r"^bebidas?\s*sin\s*alc", c, re.I):
        return "bebida"
    if re.match(r"^cigarrill", c, re.I):
        return "cigarrillos"
    return None


GRUPOS_TIENDA = ["comida", "cafe", "bebida", "cigarrillos"]
GRUPO_LABEL = {"comida": "Comida", "cafe": "Café", "bebida": "Bebida", "cigarrillos": "Cigarrillos"}

_RUBRO_RE = re.compile(r"^(\d{2}-\d{3})\s+(\S.*?)\s{2,}(-?[\d.]+,\d+)\s+(-?[\d.]+,\d+)\s*$")
_FONT_RE = re.compile(r"<FONT[^>]*>([^<]*)</FONT>", re.I)


def parse_tienda_htm(file, filename: str) -> tuple[list[dict], str | None]:
    """Lee el reporte 'Cierre de Caja' (.htm). El turno sale del nombre del
    archivo (debe terminar en TM o TT) y la fecha se busca dentro del reporte."""
    raw = file.read()
    if isinstance(raw, bytes):
        raw = raw.decode("windows-1252", errors="replace")
    lines = [m.strip() for m in _FONT_RE.findall(raw)]
    full_text = "\n".join(lines)

    base = re.sub(r"\.[^.]+$", "", filename).upper()
    turno = None
    if re.search(r"(^|[^A-Z])TM$", base):
        turno = "Mañana"
    elif re.search(r"(^|[^A-Z])TT$", base):
        turno = "Tarde"
    if not turno:
        return [], f"{filename}: no pude reconocer el turno por el nombre (debe terminar en TM o TT)."

    date_match = re.search(r"(\d{2})/(\d{2})/(\d{4})", full_text)
    if not date_match:
        return [], f"{filename}: no encontré una fecha dentro del reporte."
    dd, mm, yyyy = date_match.groups()
    fecha = f"{yyyy}-{mm}-{dd}"

    result = []
    for line in lines:
        m = _RUBRO_RE.match(line)
        if not m:
            continue
        categoria = m.group(2).strip()
        try:
            cantidad = int(m.group(3).split(",")[0].replace(".", ""))
        except ValueError:
            cantidad = 0
        importe = float(m.group(4).replace(".", "").replace(",", "."))
        result.append(
            {"fecha": fecha, "turno": turno, "categoria": categoria, "importe": importe, "cantidad": cantidad}
        )
    if not result:
        return [], f'{filename}: no encontré rubros para importar. Revisá que sea un reporte "Cierre de Caja".'
    return result, None


def parse_tienda_csv(file) -> tuple[list[dict], int]:
    df = pd.read_csv(file)
    df.columns = [c.strip().lower() for c in df.columns]
    result, errors = [], 0
    for _, r in df.iterrows():
        fecha = str(r.get("fecha", "")).strip()
        turno = normalize_tienda_turno(r.get("turno", ""))
        categoria = str(r.get("categoria", r.get("producto", ""))).strip()
        try:
            importe = float(r.get("importe"))
        except (TypeError, ValueError):
            errors += 1
            continue
        try:
            cantidad = float(r.get("cantidad", 0) or 0)
        except (TypeError, ValueError):
            cantidad = 0
        if not fecha or not turno or not categoria:
            errors += 1
            continue
        result.append({"fecha": fecha, "turno": turno, "categoria": categoria, "importe": importe, "cantidad": cantidad})
    return result, errors


# ============================================================================
# APP
# ============================================================================

from datetime import date

CLAVE_BORRADO = "Ingreso01"
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
YEARS = [2025, 2026]
YEAR_COLORS = {2025: "#3FA7D6", 2026: "#4C5FAE"}
PRODUCT_COLORS = {
    "Nafta Súper": "#E4572E",
    "Infinia": "#F2C14E",
    "Diesel 500": "#2E9E6B",
    "Infinia Diesel": "#3FA7D6",
}

st.set_page_config(page_title="Ventas de Combustible — FARMEX", page_icon="⛽", layout="wide")

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
st.markdown(
    """
    <style>
    .stApp { background-color: #0E1B2A; }
    h1, h2, h3, h4 { font-family: 'Oswald', sans-serif; letter-spacing: 0.5px; }
    .pump-number {
        font-family: 'JetBrains Mono', monospace; font-size: 2.1rem; font-weight: 700;
    }
    .pump-label { font-size: 0.75rem; color: #9FB1C6; text-transform: uppercase; letter-spacing: 1px; }
    .badge-up { color: #2E9E6B; font-weight: 700; }
    .badge-down { color: #E4572E; font-weight: 700; }
    .footer-credit { text-align: center; color: #9FB1C6; font-size: 0.75rem; padding: 24px 0 8px; }
    </style>
    """,
    unsafe_allow_html=True,
)


# ---------------------------------------------------------------------------
# Helpers de agregación — Combustible
# ---------------------------------------------------------------------------
def with_year_month(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.assign(year=[], month=[])
    out = df.copy()
    out["year"] = pd.to_datetime(out["fecha"]).dt.year
    out["month"] = pd.to_datetime(out["fecha"]).dt.month - 1
    return out


def year_totals_litros(df: pd.DataFrame) -> dict:
    t = {y: 0.0 for y in YEARS}
    if not df.empty:
        for y, v in df.groupby("year")["litros"].sum().items():
            if y in t:
                t[y] = v
    return t


def year_comparison(totals: dict):
    a, b = totals[2025], totals[2026]
    if not a and not b:
        return None
    if b >= a:
        pct = ((b - a) / a * 100) if a else 100.0
        return {"winner": 2026, "loser": 2025, "pct": pct}
    pct = ((a - b) / b * 100) if b else 100.0
    return {"winner": 2025, "loser": 2026, "pct": pct}


def monthly_table(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for i, mes in enumerate(MESES):
        row = {"mes": mes}
        for y in YEARS:
            sub = df[(df["year"] == y) & (df["month"] == i)] if not df.empty else df
            row[str(y)] = sub["litros"].sum() if not sub.empty else 0
        rows.append(row)
    return pd.DataFrame(rows)


def by_dim_table(df: pd.DataFrame, dim: str, categories: list, totals: dict) -> pd.DataFrame:
    rows = []
    for cat in categories:
        row = {dim: cat}
        for y in YEARS:
            sub = df[(df[dim] == cat) & (df["year"] == y)] if not df.empty else df
            val = sub["litros"].sum() if not sub.empty else 0
            row[str(y)] = val
            row[f"{y}pct"] = (val / totals[y] * 100) if totals[y] else 0
        rows.append(row)
    return pd.DataFrame(rows)


def pct_change(row) -> float | None:
    v25, v26 = row.get("2025", 0), row.get("2026", 0)
    if v25 > 0:
        return (v26 - v25) / v25 * 100
    if v26 > 0:
        return 100.0
    return None


def month_projection(df: pd.DataFrame, forced_month: int | None):
    if df.empty:
        return None
    if forced_month is not None:
        m = forced_month
        years_with_month = df[df["month"] == m]["year"].unique()
        y = int(max(years_with_month)) if len(years_with_month) else int(df["year"].max())
    else:
        latest = df.loc[df["fecha"].idxmax()]
        y, m = int(latest["year"]), int(latest["month"])
    month_df = df[(df["year"] == y) & (df["month"] == m)]
    days_with_data = month_df["fecha"].nunique()
    if not days_with_data:
        return None
    total_so_far = month_df["litros"].sum()
    avg_per_day = total_so_far / days_with_data
    days_in_month = pd.Period(f"{y}-{m + 1:02d}").days_in_month
    por_producto = []
    for p in PRODUCTOS:
        so_far = month_df[month_df["producto"] == p]["litros"].sum()
        avg = so_far / days_with_data
        por_producto.append({"producto": p, "so_far": so_far, "projected": avg * days_in_month})
    return {
        "year": y, "month": m, "days_with_data": days_with_data, "days_in_month": days_in_month,
        "total_so_far": total_so_far, "avg_per_day": avg_per_day, "projected": avg_per_day * days_in_month,
        "por_producto": por_producto,
    }


# ---------------------------------------------------------------------------
# Helpers — Tienda
# ---------------------------------------------------------------------------
def tienda_with_year_month_grupo(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.assign(year=[], month=[], grupo=[])
    out = df.copy()
    out["year"] = pd.to_datetime(out["fecha"]).dt.year
    out["month"] = pd.to_datetime(out["fecha"]).dt.month - 1
    out["grupo"] = out["categoria"].apply(categoria_grupo)
    return out


def tienda_year_totals(df: pd.DataFrame) -> dict:
    t = {y: dict({g: 0.0 for g in GRUPOS_TIENDA}, total=0.0) for y in YEARS}
    if not df.empty:
        grouped = df.dropna(subset=["grupo"]).groupby(["year", "grupo"])["cantidad"].sum()
        for (y, g), v in grouped.items():
            if y in t:
                t[y][g] = v
                t[y]["total"] += v
    return t


def fmt_num(n) -> str:
    try:
        return f"{int(round(n)):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "0"


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
st.markdown("## ⛽ VENTAS — FARMEX")
st.caption("Estación de servicio — comparativo 2025 / 2026")

section = st.tabs(["⛽ Combustible", "🛒 Tienda Full"])

# ===========================================================================
# COMBUSTIBLE
# ===========================================================================
with section[0]:
    cargar_tab, comp_tab = st.tabs(["Cargar", "Comparativo"])

    # ---- CARGAR ----
    with cargar_tab:
        st.subheader("📤 Carga diaria (planilla o CSV)")
        st.caption(
            "Subí la planilla del aforador tal cual la generás (.xlsx) o un CSV con columnas "
            "`fecha, turno, producto, litros, importe`. Los días y turnos ya cargados se detectan "
            "solos y no se duplican."
        )
        files = st.file_uploader(
            "Elegí uno o varios archivos", type=["xlsx", "csv"], accept_multiple_files=True, key="comb_files"
        )
        if files:
            total_added, total_skipped, total_errors = 0, 0, 0
            for f in files:
                if f.name.lower().endswith(".xlsx"):
                    rows = parse_combustible_xlsx(f)
                    errors = 0
                else:
                    rows, errors = parse_combustible_csv(f)
                added, skipped = insert_combustible(rows)
                total_added += added
                total_skipped += skipped
                total_errors += errors
            st.success(
                f"{total_added} registros nuevos agregados · {total_skipped} ya estaban cargados"
                + (f" · {total_errors} filas inválidas" if total_errors else "")
            )
            st.rerun()

        template_csv = "fecha,turno,producto,litros,importe\n2026-08-01,Mañana,Nafta Súper,320,480000\n"
        st.download_button("⬇️ Plantilla CSV", template_csv, file_name="plantilla_ventas.csv")

        with st.expander("➕ Cargar una venta manualmente"):
            with st.form("form_combustible"):
                c1, c2 = st.columns(2)
                fecha_m = c1.date_input("Fecha", value=date.today())
                turno_m = c2.selectbox("Turno", TURNOS)
                producto_m = st.selectbox("Producto", PRODUCTOS)
                c3, c4 = st.columns(2)
                litros_m = c3.number_input("Litros", min_value=0.0, step=1.0)
                importe_m = c4.number_input("Importe ($)", min_value=0.0, step=1.0)
                if st.form_submit_button("Agregar registro") and litros_m > 0:
                    insert_combustible(
                        [{"fecha": str(fecha_m), "turno": turno_m, "producto": producto_m, "litros": litros_m, "importe": importe_m}]
                    )
                    st.success("Registro agregado.")
                    st.rerun()

        st.divider()
        df_comb = load_combustible()
        st.subheader(f"Registros cargados ({len(df_comb)} en total)")

        fc1, fc2, fc3 = st.columns([1, 1, 2])
        filter_year = fc1.selectbox("Año", ["Todos"] + [str(y) for y in YEARS], key="comb_fy")
        filter_month = fc2.selectbox("Mes", ["Todos"] + MESES, key="comb_fm")

        filtered = df_comb.copy()
        if not filtered.empty:
            filtered["year"] = pd.to_datetime(filtered["fecha"]).dt.year.astype(str)
            filtered["month_idx"] = pd.to_datetime(filtered["fecha"]).dt.month - 1
            if filter_year != "Todos":
                filtered = filtered[filtered["year"] == filter_year]
            if filter_month != "Todos":
                filtered = filtered[filtered["month_idx"] == MESES.index(filter_month)]

        cdel1, cdel2 = st.columns(2)
        with cdel1:
            with st.popover("🗑️ Borrar todo"):
                pw = st.text_input("Clave", type="password", key="pw_all_comb")
                if st.button("Confirmar borrado total", key="confirm_all_comb"):
                    if pw == CLAVE_BORRADO:
                        delete_combustible_ids(df_comb["id"].tolist())
                        st.success("Borrado.")
                        st.rerun()
                    else:
                        st.error("Clave incorrecta.")
        if (filter_year != "Todos" or filter_month != "Todos") and not filtered.empty:
            with cdel2:
                with st.popover(f"🗑️ Borrar {len(filtered)} de este filtro"):
                    pw2 = st.text_input("Clave", type="password", key="pw_filt_comb")
                    if st.button("Confirmar borrado del filtro", key="confirm_filt_comb"):
                        if pw2 == CLAVE_BORRADO:
                            delete_combustible_ids(filtered["id"].tolist())
                            st.success("Borrado.")
                            st.rerun()
                        else:
                            st.error("Clave incorrecta.")

        if filtered.empty:
            st.info("No hay registros para ese filtro." if not df_comb.empty else "Todavía no cargaste ninguna venta.")
        else:
            for fecha, day_df in filtered.groupby("fecha", sort=False):
                total_l = day_df["litros"].sum()
                with st.expander(f"{fecha} — {fmt_num(total_l)} L"):
                    for turno, t_df in day_df.groupby("turno"):
                        st.markdown(f"**{turno}** — {fmt_num(t_df['litros'].sum())} L")

    # ---- COMPARATIVO ----
    with comp_tab:
        df_comb_all = with_year_month(load_combustible())
        sel_month_label = st.selectbox("Ver mes", ["Todos los meses"] + MESES, key="comb_sel_month")
        sel_month = None if sel_month_label == "Todos los meses" else MESES.index(sel_month_label)
        scoped = df_comb_all if sel_month is None else df_comb_all[df_comb_all["month"] == sel_month]

        totals = year_totals_litros(scoped)
        comparison = year_comparison(totals)

        c1, c2 = st.columns(2)
        for col, y in zip([c1, c2], YEARS):
            with col:
                st.markdown(f"**AÑO {y}**")
                if comparison and comparison["loser"] == y:
                    st.markdown(
                        f'<span class="badge-down">▼ Vendió menos (-{comparison["pct"]:.1f}%)</span>',
                        unsafe_allow_html=True,
                    )
                st.markdown(
                    f'<div class="pump-number" style="color:{YEAR_COLORS[y]}">{fmt_num(totals[y])}</div>'
                    f'<div class="pump-label">Litros vendidos</div>',
                    unsafe_allow_html=True,
                )

        proj = month_projection(df_comb_all, sel_month)
        if sel_month is not None and proj is None:
            st.info(f"Todavía no hay días cargados de {MESES[sel_month]} para proyectar.")
        elif proj:
            st.markdown(f"### 📈 Proyección — {MESES[proj['month']]} {proj['year']}")
            st.caption(
                f"Con {proj['days_with_data']} día(s) cargado(s) de {proj['days_in_month']} del mes, "
                "a este ritmo el mes cerraría en:"
            )
            pc1, pc2, pc3 = st.columns(3)
            pc1.metric("Litros hasta ahora", fmt_num(proj["total_so_far"]))
            pc2.metric("Promedio diario", fmt_num(proj["avg_per_day"]))
            pc3.metric("Proyección fin de mes", fmt_num(proj["projected"]))
            proj_df = pd.DataFrame(proj["por_producto"])
            proj_df["Hasta ahora"] = proj_df["so_far"].apply(fmt_num) + " L"
            proj_df["Proyección"] = proj_df["projected"].apply(fmt_num) + " L"
            st.dataframe(
                proj_df[["producto", "Hasta ahora", "Proyección"]].rename(columns={"producto": "Producto"}),
                hide_index=True, use_container_width=True,
            )

        st.markdown("### Venta total por mes — 2025 vs 2026 (litros)")
        monthly = monthly_table(df_comb_all)
        monthly_long = monthly.melt(id_vars="mes", var_name="Año", value_name="Litros")
        fig = px.bar(
            monthly_long, x="mes", y="Litros", color="Año", barmode="group",
            color_discrete_map={"2025": YEAR_COLORS[2025], "2026": YEAR_COLORS[2026]},
            category_orders={"mes": MESES},
        )
        fig.update_layout(paper_bgcolor="#152840", plot_bgcolor="#152840", font_color="#F4EFE6")
        st.plotly_chart(fig, use_container_width=True)

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("### Venta por producto")
            bp = by_dim_table(scoped, "producto", PRODUCTOS, totals)
            bp_long = bp.melt(id_vars="producto", value_vars=["2025", "2026"], var_name="Año", value_name="Litros")
            fig2 = px.bar(
                bp_long, x="Litros", y="producto", color="Año", orientation="h", barmode="group",
                color_discrete_map={"2025": YEAR_COLORS[2025], "2026": YEAR_COLORS[2026]},
            )
            fig2.update_layout(paper_bgcolor="#152840", plot_bgcolor="#152840", font_color="#F4EFE6")
            st.plotly_chart(fig2, use_container_width=True)

        with col_b:
            st.markdown("### Venta por turno")
            bt = by_dim_table(scoped, "turno", TURNOS, totals)
            for _, row in bt.iterrows():
                pct = pct_change(row)
                if pct is None:
                    st.caption(f"{row['turno']}: sin datos")
                else:
                    cls = "badge-up" if pct >= 0 else "badge-down"
                    arrow = "▲" if pct >= 0 else "▼"
                    st.markdown(f'{row["turno"]}: <span class="{cls}">{arrow} {pct:+.1f}%</span>', unsafe_allow_html=True)
            bt_long = bt.melt(id_vars="turno", value_vars=["2025", "2026"], var_name="Año", value_name="Litros")
            fig3 = px.bar(
                bt_long, x="turno", y="Litros", color="Año", barmode="group",
                color_discrete_map={"2025": YEAR_COLORS[2025], "2026": YEAR_COLORS[2026]},
            )
            fig3.update_layout(paper_bgcolor="#152840", plot_bgcolor="#152840", font_color="#F4EFE6")
            st.plotly_chart(fig3, use_container_width=True)

        st.markdown("### Mix de productos")
        mix_rows = []
        for p in PRODUCTOS:
            row = bp[bp["producto"] == p].iloc[0]
            mix_rows.append(
                {"Producto": p, "% 2025": f"{row['2025pct']:.0f}%", "L 2025": fmt_num(row["2025"]),
                 "% 2026": f"{row['2026pct']:.0f}%", "L 2026": fmt_num(row["2026"])}
            )
        st.dataframe(pd.DataFrame(mix_rows), hide_index=True, use_container_width=True)

    st.markdown('<div class="footer-credit">Desarrollado por Lucas Sellecchia</div>', unsafe_allow_html=True)

# ===========================================================================
# TIENDA FULL
# ===========================================================================
with section[1]:
    cargar_tab_t, comp_tab_t = st.tabs(["Cargar", "Comparativo"])

    with cargar_tab_t:
        st.subheader("📤 Carga diaria — Tienda Full")
        st.caption(
            "Subí el reporte de cierre de caja tal cual lo generás (.htm) — el nombre debe terminar en "
            "`TM` o `TT` según el turno (ej: `31-07__TM.htm`). También podés subir un CSV con columnas "
            "`fecha, turno, categoria, importe, cantidad`."
        )
        files_t = st.file_uploader(
            "Elegí uno o varios archivos", type=["htm", "html", "csv"], accept_multiple_files=True, key="tienda_files"
        )
        if files_t:
            total_added, total_skipped = 0, 0
            msgs = []
            for f in files_t:
                if f.name.lower().endswith(".csv"):
                    rows, errors = parse_tienda_csv(f)
                    if errors:
                        msgs.append(f"{f.name}: {errors} filas inválidas")
                else:
                    rows, err_msg = parse_tienda_htm(f, f.name)
                    if err_msg:
                        msgs.append(err_msg)
                added, skipped = insert_tienda(rows)
                total_added += added
                total_skipped += skipped
            st.success(f"{total_added} registros nuevos agregados · {total_skipped} ya estaban cargados")
            for m in msgs:
                st.warning(m)
            st.rerun()

        template_csv_t = "fecha,turno,categoria,importe,cantidad\n2026-08-01,Mañana,Kiosco,45000,10\n"
        st.download_button("⬇️ Plantilla CSV", template_csv_t, file_name="plantilla_tienda.csv")

        with st.expander("➕ Cargar una venta manualmente"):
            with st.form("form_tienda"):
                c1, c2 = st.columns(2)
                fecha_t = c1.date_input("Fecha", value=date.today(), key="fecha_t")
                turno_t = c2.selectbox("Turno", TURNOS_TIENDA, key="turno_t")
                categoria_t = st.text_input("Categoría / producto", placeholder="Ej: Kiosco, Café, Panadería…")
                importe_t = st.number_input("Importe ($)", min_value=0.0, step=1.0, key="importe_t")
                if st.form_submit_button("Agregar registro") and categoria_t and importe_t > 0:
                    insert_tienda(
                        [{"fecha": str(fecha_t), "turno": turno_t, "categoria": categoria_t, "importe": importe_t, "cantidad": 0}]
                    )
                    st.success("Registro agregado.")
                    st.rerun()

        st.divider()
        df_tienda = load_tienda()
        st.subheader(f"Registros cargados ({len(df_tienda)} en total)")

        ft1, ft2 = st.columns(2)
        filter_year_t = ft1.selectbox("Año", ["Todos"] + [str(y) for y in YEARS], key="tienda_fy")
        filter_month_t = ft2.selectbox("Mes", ["Todos"] + MESES, key="tienda_fm")

        filtered_t = df_tienda.copy()
        if not filtered_t.empty:
            filtered_t["year"] = pd.to_datetime(filtered_t["fecha"]).dt.year.astype(str)
            filtered_t["month_idx"] = pd.to_datetime(filtered_t["fecha"]).dt.month - 1
            filtered_t["grupo"] = filtered_t["categoria"].apply(categoria_grupo)
            if filter_year_t != "Todos":
                filtered_t = filtered_t[filtered_t["year"] == filter_year_t]
            if filter_month_t != "Todos":
                filtered_t = filtered_t[filtered_t["month_idx"] == MESES.index(filter_month_t)]

        cdel1t, cdel2t = st.columns(2)
        with cdel1t:
            with st.popover("🗑️ Borrar todo"):
                pw_t = st.text_input("Clave", type="password", key="pw_all_tienda")
                if st.button("Confirmar borrado total", key="confirm_all_tienda"):
                    if pw_t == CLAVE_BORRADO:
                        delete_tienda_ids(df_tienda["id"].tolist())
                        st.success("Borrado.")
                        st.rerun()
                    else:
                        st.error("Clave incorrecta.")
        if (filter_year_t != "Todos" or filter_month_t != "Todos") and not filtered_t.empty:
            with cdel2t:
                with st.popover(f"🗑️ Borrar {len(filtered_t)} de este filtro"):
                    pw2t = st.text_input("Clave", type="password", key="pw_filt_tienda")
                    if st.button("Confirmar borrado del filtro", key="confirm_filt_tienda"):
                        if pw2t == CLAVE_BORRADO:
                            delete_tienda_ids(filtered_t["id"].tolist())
                            st.success("Borrado.")
                            st.rerun()
                        else:
                            st.error("Clave incorrecta.")

        if filtered_t.empty:
            st.info("No hay registros para ese filtro." if not df_tienda.empty else "Todavía no cargaste ninguna venta de Tienda Full.")
        else:
            for fecha, day_df in filtered_t.groupby("fecha", sort=False):
                with st.expander(f"{fecha}"):
                    for turno, t_df in day_df.groupby("turno"):
                        counts = t_df.groupby("grupo")["cantidad"].sum()
                        linea = " · ".join(
                            f"{GRUPO_LABEL[g]}: **{fmt_num(counts.get(g, 0))}**" for g in GRUPOS_TIENDA
                        )
                        st.markdown(f"**{turno}** — {linea}")

    with comp_tab_t:
        df_tienda_all = tienda_with_year_month_grupo(load_tienda())
        sel_month_t_label = st.selectbox("Ver mes", ["Todos los meses"] + MESES, key="tienda_sel_month")
        sel_month_t = None if sel_month_t_label == "Todos los meses" else MESES.index(sel_month_t_label)
        scoped_t = df_tienda_all if sel_month_t is None else df_tienda_all[df_tienda_all["month"] == sel_month_t]

        totals_t = tienda_year_totals(scoped_t)
        a25, a26 = totals_t[2025]["total"], totals_t[2026]["total"]
        comparison_t = None
        if a25 or a26:
            if a26 >= a25:
                comparison_t = {"winner": 2026, "loser": 2025, "pct": ((a26 - a25) / a25 * 100) if a25 else 100.0}
            else:
                comparison_t = {"winner": 2025, "loser": 2026, "pct": ((a25 - a26) / a26 * 100) if a26 else 100.0}

        c1t, c2t = st.columns(2)
        for col, y in zip([c1t, c2t], YEARS):
            with col:
                st.markdown(f"**AÑO {y}**")
                if comparison_t and comparison_t["loser"] == y:
                    st.markdown(
                        f'<span class="badge-down">▼ Vendió menos (-{comparison_t["pct"]:.1f}%)</span>',
                        unsafe_allow_html=True,
                    )
                mcols = st.columns(4)
                for mc, g in zip(mcols, GRUPOS_TIENDA):
                    mc.markdown(
                        f'<div class="pump-number" style="color:{YEAR_COLORS[y]};font-size:1.4rem">{fmt_num(totals_t[y][g])}</div>'
                        f'<div class="pump-label">{GRUPO_LABEL[g]}</div>',
                        unsafe_allow_html=True,
                    )

        st.markdown("### Unidades vendidas por mes — 2025 vs 2026")
        rows_m = []
        for i, mes in enumerate(MESES):
            row = {"mes": mes}
            for y in YEARS:
                sub = df_tienda_all[(df_tienda_all["year"] == y) & (df_tienda_all["month"] == i) & df_tienda_all["grupo"].notna()]
                row[str(y)] = sub["cantidad"].sum() if not sub.empty else 0
            rows_m.append(row)
        monthly_t = pd.DataFrame(rows_m)
        monthly_t_long = monthly_t.melt(id_vars="mes", var_name="Año", value_name="Unidades")
        figt = px.bar(
            monthly_t_long, x="mes", y="Unidades", color="Año", barmode="group",
            color_discrete_map={"2025": YEAR_COLORS[2025], "2026": YEAR_COLORS[2026]},
            category_orders={"mes": MESES},
        )
        figt.update_layout(paper_bgcolor="#152840", plot_bgcolor="#152840", font_color="#F4EFE6")
        st.plotly_chart(figt, use_container_width=True)

        st.markdown("### Unidades por turno")
        rows_turno = []
        for t in TURNOS_TIENDA:
            row = {"turno": t}
            for y in YEARS:
                sub = scoped_t[(scoped_t["turno"] == t) & (scoped_t["year"] == y) & scoped_t["grupo"].notna()]
                val = sub["cantidad"].sum() if not sub.empty else 0
                row[str(y)] = val
            rows_turno.append(row)
        bt_t = pd.DataFrame(rows_turno)
        for _, row in bt_t.iterrows():
            pct = pct_change(row)
            if pct is None:
                st.caption(f"{row['turno']}: sin datos")
            else:
                cls = "badge-up" if pct >= 0 else "badge-down"
                arrow = "▲" if pct >= 0 else "▼"
                st.markdown(f'{row["turno"]}: <span class="{cls}">{arrow} {pct:+.1f}%</span>', unsafe_allow_html=True)
        bt_t_long = bt_t.melt(id_vars="turno", value_vars=["2025", "2026"], var_name="Año", value_name="Unidades")
        figtt = px.bar(
            bt_t_long, x="turno", y="Unidades", color="Año", barmode="group",
            color_discrete_map={"2025": YEAR_COLORS[2025], "2026": YEAR_COLORS[2026]},
        )
        figtt.update_layout(paper_bgcolor="#152840", plot_bgcolor="#152840", font_color="#F4EFE6")
        st.plotly_chart(figtt, use_container_width=True)

        st.markdown("### Versus por rubro — 2025 vs 2026")
        rubro_rows = []
        for g in GRUPOS_TIENDA:
            v25, v26 = totals_t[2025][g], totals_t[2026][g]
            p25 = (v25 / totals_t[2025]["total"] * 100) if totals_t[2025]["total"] else 0
            p26 = (v26 / totals_t[2026]["total"] * 100) if totals_t[2026]["total"] else 0
            rubro_rows.append(
                {"Rubro": GRUPO_LABEL[g], "% 2025": f"{p25:.0f}%", "u. 2025": fmt_num(v25),
                 "% 2026": f"{p26:.0f}%", "u. 2026": fmt_num(v26)}
            )
        st.dataframe(pd.DataFrame(rubro_rows), hide_index=True, use_container_width=True)
        if totals_t[2025]["total"] == 0:
            st.caption(
                "Todavía no hay datos de 2025 para comparar — en cuanto subas los reportes del año "
                "anterior, este cuadro arma el versus completo por rubro."
            )

    st.markdown('<div class="footer-credit">Desarrollado por Lucas Sellecchia</div>', unsafe_allow_html=True)
